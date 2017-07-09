import os, struct
import openpyxl as opxl
from openpyxl.drawing.image import Image
import matplotlib as plt
from array import array as pyarray
from numpy import append, array, int8, uint8, zeros as np

positive_class = 8
negative_class = 2
excel_file = "Assignment_3_ Submission_Template.xlsx"
bin_size = 25


def write_excel_data(x, sheet_name, start_row, start_col):
    from pandas import DataFrame, ExcelWriter
    from openpyxl import load_workbook
    df=DataFrame(x)
    book = load_workbook(excel_file)
    writer = ExcelWriter(excel_file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.to_excel(writer, sheet_name=sheet_name,startrow=start_row-1, startcol=start_col-1, header=False, index=False)
    writer.save()
    writer.close()


def write_result(result, xls_row, xls_col) :
    write_excel_data(result, "Results", xls_row, xls_col)

write_result([["Abhishek Chaturvedi"]],1,1)


def load_mnist(dataset="training", digits=range(10),
               path='mnist_data'):
    """
    Adapted from: http://cvxopt.org/applications/svm/index.html?highlight=mnist
    """

    if dataset == "training":
        fname_img = os.path.join(path, 'train-images-idx3-ubyte')
        fname_lbl = os.path.join(path, 'train-labels-idx1-ubyte')
    elif dataset == "testing":
        fname_img = os.path.join(path, 't10k-images-idx')
        fname_lbl = os.path.join(path, 't10k-labels-idx')
    else:
        raise ValueError("dataset must be 'testing' or 'training'")

    flbl = open(fname_lbl, 'rb')
    magic_nr, size = struct.unpack(">II", flbl.read(8))
    lbl = pyarray("b", flbl.read())
    flbl.close()

    fimg = open(fname_img, 'rb')
    magic_nr, size, rows, cols = struct.unpack(">IIII", fimg.read(16))
    img = pyarray("B", fimg.read())
    fimg.close()

    ind = [k for k in range(size) if lbl[k] in digits]
    N = len(ind)

    images = zeros((N, rows, cols), dtype=uint8)
    labels = zeros((N, 1), dtype=int8)
    for i in range(len(ind)):
        images[i] = array(img[ind[i] * rows * cols: (ind[i] + 1) * rows * cols]).reshape((rows, cols))
        labels[i] = lbl[ind[i]]

    return images, labels


def create_hist(input_data, max_1, min_1, slot_1, max_2, min_2, slot_2):
    i = 0
    hist = np.zeros((bin_size, bin_size))
    # print(hist)
    for h in np.arange(min_1, max_1, slot_1):
        j = 0
        for s in np.arange(min_2, max_2, slot_2):
            temp_data = input_data[(np.where(np.logical_and(h <= input_data[:, 0], input_data[:, 0] < h + slot_1)))]
            hist[i][j] = temp_data[(np.where(np.logical_and(s <= temp_data[:, 1], temp_data[:, 1] < s + slot_2)))][:,
                         0].size
            j += 1
        i += 1
    return hist


from pylab import *
from numpy import *

images, labels = load_mnist('training', digits=[2,8])

n_values = np.flip(np.unique(labels, return_counts=True)[1], axis=0)
#print(n_values)
#writeExcelData(n_values,"Results",6,2)

# converting from NX28X28 array into NX784 array
flatimages = list()
for i in images:
    flatimages.append(i.ravel())
X = np.asarray(flatimages)


print("Check shape of matrix", X.shape)
print("Check Mins and Max Values",np.amin(X),np.amax(X))
#print("\nCheck training vector by plotting image", labels[21])
# plt.imshow(X[21].reshape(28, 28),interpolation='None', cmap=cm.gray)
# show()

mu=X.mean(dtype=np.int32, axis=0)
write_result([mu],2,2)

Z=X-mu
C=np.cov(Z,rowvar=False)
[L,V]=np.linalg.eigh(C);
L=np.flipud(L);
V=np.flipud(V.T);
P=np.dot(Z,V.T)
#R=np.dot(P,V);
#print(R)
write_result(V[0:2,:],3,2)
#Xrec2=(np.dot(P[:,0:2],V[0:2,:]))+mu;
#print("\n2 dimensions R", Xrec2[21])

training_data_x2 = P[:,0:2]
i_8 = [k for k in range(labels.size) if labels[k] == positive_class]
i_2 = [k for k in range(labels.size) if labels[k] == negative_class]

training_data_8 = training_data_x2[i_8, :]
mu_8=training_data_8.mean(dtype=np.float32, axis=0)
write_result([mu_8],9,2)

training_data_2 = training_data_x2[i_2, :]
mu_2=training_data_2.mean(dtype=np.float32, axis=0)
write_result([mu_2],10,2)

cov_8 = np.cov(training_data_8, rowvar=False)
write_result(cov_8,12,2)

cov_2 = np.cov(training_data_2, rowvar=False)
write_result(cov_2,14,2)

hist_max = np.amax(training_data_x2, axis=0)
hist_min = np.amin(training_data_x2, axis=0)
write_result(np.array([hist_min, hist_max]).T,17,2)

pc1_slot = (hist_max[0] - hist_min[0]) / bin_size
pc2_slot = (hist_max[1] - hist_min[1]) / bin_size

hist_8 = create_hist(training_data_8, hist_max[0], hist_min[0], pc1_slot, hist_max[1], hist_min[1], pc2_slot)
write_result(hist_8,20,2)

hist_2 = create_hist(training_data_2, hist_max[0], hist_min[0], pc1_slot, hist_max[1], hist_min[1], pc2_slot)
write_result(hist_2,46,2)


def query_vector(row_num, xls_row_num, xls_row, write_xls):
    #print(labels[row_num])
    if xls_row>0 :
        write_result([labels[row_num]], xls_row, 2)
    xp = X[row_num]
    zp = Z[row_num]
    pp = training_data_x2[row_num]
    rp = np.dot(pp,V[0:2,:])
    xrecp = rp + mu
    #plt.imshow(xrecp.reshape(28, 28),interpolation='None', cmap=cm.gray)
    #show()
    if write_xls :
        write_result([xp], xls_row_num, 2)
        write_result([zp], xls_row_num + 1, 2)
        write_result([pp], xls_row_num + 2, 2)
        write_result([rp], xls_row_num + 3, 2)
        write_result([xrecp], xls_row_num + 4, 2)
    return pp


def query_result_hist(query):
    i = int(round((query[0] - hist_min[0])/pc1_slot)) - 1
    j = int(round((query[1] - hist_min[1])/pc2_slot)) - 1
    prob = hist_8[i][j] / (hist_8[i][j] + hist_2[i][j])
    #print ("Hist probability of being 8 : ", i, j, hist_8[i][j], hist_2[i][j], prob)
    return round(prob,4)


def calc_gaussian_pd(x, mean, cov, count):
    return np.multiply(count, np.multiply(1 / (2 * np.pi * (np.sqrt(np.linalg.det(cov)))), np.exp(np.multiply(-1 / 2,
                                                                                                              np.dot(
                                                                                                                  np.dot(
                                                                                                                      np.subtract(
                                                                                                                          x,
                                                                                                                          mean),
                                                                                                                      np.linalg.inv(
                                                                                                                          cov)),
                                                                                                                  np.transpose(
                                                                                                                      np.subtract(
                                                                                                                          x,
                                                                                                                          mean)))))))


def query_result_bayesian(x):
    n8 = n_values[0]
    n2 = n_values[1]

    prob = np.divide(calc_gaussian_pd(x, mu_8, cov_8, n8),
                    np.add(calc_gaussian_pd(x, mu_8, cov_8, n8),
                           calc_gaussian_pd(x, mu_2, cov_2, n2)))
    #print("Bayes probability of being 8: ", prob)
    return round(prob,4)


def find_training_accuracy():
    hist_hit = 0
    bayes_hit = 0
    for q in range(labels.size):
    #for q in range(5000):
        hist_prob = query_result_hist(query_vector(q, 0, 0, False))
        hist_hit += 1 if ((labels[q] ==8 and hist_prob > 0.5) or (labels[q] == 2 and hist_prob < 0.5)) else 0
        bayes_prob = query_result_bayesian(query_vector(q, 0, 0, False))
        bayes_hit += 1 if ((labels[q] == 8 and bayes_prob > 0.5) or (labels[q] == 2 and bayes_prob < 0.5)) else 0
        #print(" Vector : ", q, labels[q], hist_prob, bayes_prob)

    hist_success_rate = round((hist_hit/labels.size)*100, 2)
    write_result([hist_success_rate], 97, 2)
    print("hist_success_rate ", hist_success_rate)
    bayes_success_rate = round((bayes_hit/labels.size)*100, 2)
    write_result([bayes_success_rate], 98, 2)
    print("bayes_success_rate ", bayes_success_rate)


prob_8 = query_result_hist(query_vector(6, 74, 88, True))
write_result([prob_8], 89, 3)
write_result([8 if prob_8>0.5 else 2], 89, 2)

prob_8 = query_result_bayesian(query_vector(6, 74, 88, False))
write_result([prob_8], 90, 3)
write_result([8 if prob_8>0.5 else 2], 90, 2)

prob_8 = query_result_hist(query_vector(9, 80, 92, True))
write_result([prob_8], 93, 3)
write_result([8 if prob_8>0.5 else 2], 93, 2)

prob_8 = query_result_bayesian(query_vector(9, 80, 92, False))
write_result([prob_8], 94, 3)
write_result([8 if prob_8>0.5 else 2], 94, 2)

find_training_accuracy()

plt.scatter(training_data_8[:,0], training_data_8[:,1], marker='.', color="g")
plt.scatter(training_data_2[:,0], training_data_2[:,1], marker='x', color="r")
plt.legend("8 and 2 scatter plot")
matplotlib.pyplot.savefig('scatter-plot.png')

wb = opxl.load_workbook(excel_file)
ws = wb.worksheets[1]
img = Image('scatter-plot.png')
ws.add_image(img, 'A1')
wb.save(excel_file)

